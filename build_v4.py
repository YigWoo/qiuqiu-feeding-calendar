# -*- coding: utf-8 -*-
from pathlib import Path
from copy import deepcopy
from datetime import datetime
from html import escape
import json, re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
OLD_HTML = ROOT.parent / 'ball-feeding-calendar' / '球球辅食日历.html'
OUT_HTML = ROOT / 'index.html'
OUT_XLSX = ROOT / '球球辅食逐日计划_审查优化V4_口感安全与网页版.xlsx'

old = OLD_HTML.read_text(encoding='utf-8')
plan = json.loads(re.search(r'const PLAN = (.*?);\nconst RULES', old, re.S).group(1))
rules = json.loads(re.search(r'const RULES = (.*?);\nconst KEY', old, re.S).group(1))

# Revisions keep each meal's planned amount intact. They clarify safe texture and serving.
def revise_text(text, day):
    if not text or text == '—':
        return text
    if day <= 3:
        text = text.replace('高铁燕麦/多谷物婴儿米粉糊', '单一配方高铁婴儿米粉糊')
    if day <= 95:
        text = text.replace('全熟鸡蛋碎', '全熟鸡蛋压成细末')
        text = text.replace('嫩豆腐碎', '嫩豆腐压成细碎泥')
    text = text.replace('软饭/软面/稠粥', '软饭、软面或稠粥（三选一）')
    return text

def suggestion(row, day):
    meal = ' '.join(str(row.get(k) or '') for k in ('早餐','午餐','晚餐'))
    notice = (str(row.get('新增食材或观察') or '') + ' ' + str(row.get('油脂/过敏原/其他') or '')).strip()
    flavor = []
    safety = []
    if day <= 3:
        flavor.append('起步只用一种单一配方高铁米粉；调成能挂勺的浓稠糊，不做稀汤。')
    elif day <= 34:
        flavor.append('咸口食材可压成细泥，温热不烫再喂；一次只做小半碗，吃完再添。')
    elif day <= 95:
        flavor.append('咸口和甜口分勺/分小碗，不混成一碗；鸡蛋、豆腐压细后拌入熟食。')
    else:
        flavor.append('每餐主食只选一种；软饭、软面或稠粥任选其一，保持湿润软烂。')
    if any(x in meal for x in ('鱼','鳕','虾')):
        safety.append('鱼逐口复核无刺；虾彻底煮熟、去壳去硬壳。')
    if any(x in meal for x in ('花生','核桃','芝麻')):
        safety.append('坚果/芝麻酱必须调稀，绝不整粒、整勺或干粉直接喂。')
    if '首次' in notice or '新增' in notice:
        safety.append('今天有新食材：白天吃、精神状态好时吃；当天不再新增其他食材，观察皮疹、持续呕吐、喘鸣等。')
    if not safety:
        safety.append('坐稳、全程看护；不强喂。拒绝、转头或明显疲劳就停止。')
    return ' '.join(flavor), ' '.join(safety)

revised = deepcopy(plan)
for i, row in enumerate(revised, start=1):
    for col in ('早餐','午餐','晚餐'):
        row[col] = revise_text(row[col], i)
    row['口感与喂食建议'], row['安全重点'] = suggestion(row, i)

# V4 workbook: reconstructed from the approved V3 payload because the original cache is no longer available.
wb = Workbook()
ws = wb.active
ws.title = '逐日计划'
headers = ['日期','月龄阶段','餐次','早餐','午餐','晚餐','当日目标总量','质地/进阶','新增食材或观察','油脂/过敏原/其他','奶与喂养提示','口感与喂食建议','安全重点']
ws.append(headers)
for row in revised:
    ws.append([datetime.strptime(row['日期'], '%Y-%m-%d') if k == '日期' else row.get(k,'') for k in headers])
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='39765B')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
for i, width in enumerate([14,16,10,38,48,48,18,28,36,40,34,44,48], 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for r in range(2, ws.max_row+1): ws.row_dimensions[r].height = 54

rs = wb.create_sheet('执行规则与资料')
rs.append(['模块','条目','内容','参考链接/日期'])
for item in rules:
    rs.append([item.get('模块',''),item.get('条目',''),item.get('内容',''),item.get('参考链接/日期','')])
rs.append(['V4审查优化','本次改动','在不改动原计划日期、食材总量、油量或过敏原时序的前提下：起步期统一明确为单一配方高铁米粉；6月龄早期的蛋和豆腐明确压细；三餐期的“软饭/软面/稠粥”明确为三选一；逐日新增“口感与喂食建议”和“安全重点”。原V3网页数据保留在本项目中。','2026-08-28'])
for cell in rs[1]:
    cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='39765B'); cell.alignment = Alignment(wrap_text=True)
for row in rs.iter_rows(min_row=2):
    for cell in row: cell.alignment = Alignment(vertical='top',wrap_text=True)
for i,w in enumerate([16,26,90,42],1): rs.column_dimensions[get_column_letter(i)].width=w
rs.freeze_panes='A2'; rs.auto_filter.ref=rs.dimensions
wb.save(OUT_XLSX)

plan_json = json.dumps(revised, ensure_ascii=False)
rules_json = json.dumps(rules, ensure_ascii=False)
html = r'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover"><meta name="theme-color" content="#f6f8f1"><title>球球今天吃什么</title><style>
:root{--ink:#27372e;--muted:#6d7a70;--cream:#f6f8f1;--card:#fff;--green:#34745a;--mint:#e9f4e8;--amber:#fff5d6;--rose:#fff0eb;--line:#dde7dd;--shadow:0 8px 28px rgba(41,69,48,.08)}*{box-sizing:border-box}body{margin:0;background:var(--cream);color:var(--ink);font:17px/1.55 -apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif}.shell{max-width:650px;margin:auto;padding:16px 13px 55px}.top{display:flex;justify-content:space-between;align-items:center;margin:2px 0 14px}.brand{font-size:22px;font-weight:850;letter-spacing:.2px}.brand small{display:block;font-size:13px;color:var(--muted);font-weight:600}.card{background:var(--card);border:1px solid var(--line);border-radius:21px;padding:17px;margin:12px 0;box-shadow:var(--shadow)}.hero{background:linear-gradient(135deg,#e3f3e8,#fffdf3)}.date{font-size:15px;color:var(--muted);font-weight:700}.title{font-size:29px;line-height:1.18;margin:5px 0 7px;font-weight:900}.badge{display:inline-block;border-radius:999px;background:var(--green);color:#fff;padding:4px 11px;font-weight:800;font-size:14px}.status{margin-top:11px;color:#4f6156;font-size:15px}.meal{padding:15px 0;border-bottom:1px solid var(--line)}.meal:last-child{border-bottom:0}.meal h2{font-size:20px;margin:0 0 10px}.foodGrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:9px}.foodItem{min-width:0;border:1px solid var(--line);border-radius:16px;background:#fbfdf9;padding:9px;display:flex;gap:9px;align-items:center}.foodPic{flex:0 0 58px;width:58px;height:58px;border-radius:14px;overflow:hidden;background:#f3f5ef}.foodPic img{display:block;width:58px;height:58px}.foodName{font-weight:850;line-height:1.35;font-size:15px;word-break:break-word}.foodSub{font-size:12px;color:var(--muted);margin-top:3px}.mealFull{margin:10px 0 0;padding:9px 10px;border-radius:11px;background:#f5f8f4;font-size:13px;color:#526158}.food{font-weight:700;line-height:1.55}.dash{color:#93a099}.tag{font-size:13px;color:var(--muted);margin-top:5px}.notice{background:var(--amber);border-color:#f0dd93}.notice.warn{background:var(--rose);border-color:#f1c7be}.notice h3{margin:0 0 5px;font-size:17px}.notice p{margin:0}.advice{display:grid;gap:10px}.advice div{padding:12px;border-radius:14px;background:#f5f8f4}.advice b{display:block;font-size:14px;color:var(--green);margin-bottom:3px}.nav{display:grid;grid-template-columns:80px 1fr 80px;gap:8px;align-items:center}.nav button{white-space:nowrap;font-size:14px}.datepick{width:100%;font:inherit;border:1px solid var(--line);padding:10px 6px;border-radius:11px;background:#fff;text-align:center}button{font:inherit;border:0;border-radius:12px;padding:10px 12px;font-weight:800;color:#fff;background:var(--green);cursor:pointer}button.ghost{background:#ebf0eb;color:var(--ink)}button:focus,input:focus{outline:3px solid #b8e1c4;outline-offset:2px}.quick{display:flex;gap:8px;margin-top:10px}.quick button{flex:1}.help{font-size:14px;color:var(--muted);margin-top:9px}.details summary{font-weight:800;cursor:pointer;color:var(--green)}.details{font-size:15px}.hidden{display:none!important}.manager{border:2px dashed #bfdbc7}.manager label{display:block;font-size:15px;font-weight:800;margin:12px 0 5px}.manager input,.manager textarea{width:100%;font:inherit;border:1px solid var(--line);border-radius:11px;padding:10px;background:#fff}.manager textarea{min-height:64px}.row{display:flex;gap:9px;margin:10px 0;flex-wrap:wrap}.record{border-top:1px solid var(--line);padding:10px 0;font-size:15px}.pause{background:var(--rose);border-color:#edc3b9}@media(max-width:390px){.shell{padding:12px 9px 44px}.card{padding:14px}.title{font-size:26px}.nav{grid-template-columns:68px 1fr 68px}.nav button{padding:10px 4px}.foodGrid{grid-template-columns:1fr}.foodName{font-size:16px}}
</style></head><body><main class="shell"><header class="top"><div class="brand">球球今天吃什么<small>辅食日历 · 家人查看版</small></div><button class="ghost" id="openManage">设置</button></header><section class="card hero" id="hero"></section><section class="card" id="meals"></section><section class="card notice" id="observe"></section><section class="card"><div class="nav"><button class="ghost" id="prev">← 前一天</button><input id="pick" class="datepick" type="date" aria-label="选择日期"><button class="ghost" id="next">后一天 →</button></div><div class="quick"><button class="ghost" id="goToday">回到今天</button><button id="share">复制家人链接</button></div><div class="help" id="shareHint"></div></section><section class="card"><div class="advice" id="advice"></div></section><details class="card details"><summary>给家里人的安全提醒</summary><p>坐稳、全程看护，不强喂。食物做软、做碎、做湿润；整粒坚果、整颗葡萄、硬块、生食、带刺鱼和烫食都不喂。出现呼吸困难、反复呕吐、面唇舌肿、全身风团等，立即就医。</p></details><section class="card manager hidden" id="manager"><h2 style="margin-top:0">管理员设置</h2><div class="help">修改后请点“复制家人链接”发到群里；家人用这个链接打开，会看到相同的开始日期和暂停记录。</div><label>实际开始日期</label><input id="start" type="date"><button id="saveStart" style="margin-top:9px">保存开始日期</button><hr style="border:0;border-top:1px solid var(--line);margin:20px 0"><h3>当天暂停、计划顺延</h3><div class="help">暂停日不消耗计划进度，下一天继续同一份计划。</div><label>暂停日期</label><input id="pauseDate" type="date"><label>原因（选填）</label><textarea id="pauseReason" placeholder="如：接种后不适、出门"></textarea><button id="addPause" style="margin-top:9px;background:#b6594e">记录暂停并顺延</button><h3>已记录暂停</h3><div id="pauseList" class="help"></div></section></main><script>
const PLAN=__PLAN__;const RULES=__RULES__;const KEY='ball-food-viewer-v4';const $=x=>document.getElementById(x);const iso=d=>{let x=new Date(d);x.setMinutes(x.getMinutes()-x.getTimezoneOffset());return x.toISOString().slice(0,10)};const today=()=>iso(new Date());const plus=(s,n)=>{let d=new Date(s+'T12:00:00');d.setDate(d.getDate()+n);return iso(d)};const diff=(a,b)=>Math.round((new Date(b+'T12:00:00')-new Date(a+'T12:00:00'))/864e5);const b64e=s=>btoa(unescape(encodeURIComponent(s)));const b64d=s=>decodeURIComponent(escape(atob(s)));
function load(){let q=new URLSearchParams(location.search);if(q.has('start')){try{return{start:q.get('start'),pauses:JSON.parse(b64d(q.get('pauses')||'')),shared:true}}catch(e){return{start:q.get('start'),pauses:[],shared:true}}}try{return{start:'2026-09-13',pauses:[],...JSON.parse(localStorage.getItem(KEY)||'{}'),shared:false}}catch(e){return{start:'2026-09-13',pauses:[],shared:false}}}let state=load(),shown=today();function save(){if(!state.shared)localStorage.setItem(KEY,JSON.stringify({start:state.start,pauses:state.pauses}))}function ix(d){return diff(state.start,d)-state.pauses.filter(x=>x.date>=state.start&&x.date<=d).length}function paused(d){return state.pauses.find(x=>x.date===d)}function e(s){return String(s||'').replace(/[&<>\"]/g,x=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[x]))}function emoji(text){let m=[['虾','🦐'],['鱼','🐟'],['鳕','🐟'],['鸡','🍗'],['牛','🥩'],['猪','🥩'],['蛋','🥚'],['豆腐','◻️'],['花生','🥜'],['核桃','🌰'],['芝麻','⚪'],['酸奶','🥣'],['苹果','🍎'],['梨','🍐'],['香蕉','🍌'],['牛油果','🥑'],['南瓜','🎃'],['胡萝卜','🥕'],['红薯','🍠'],['面','🍜'],['饭','🍚'],['粥','🥣'],['米粉','🥣']];let a=m.find(x=>text.includes(x[0]));return a?a[1]:'🍽️'}function svg(text){let icon=emoji(text), hue=icon==='🐟'?'#dcedf3':icon==='🥕'||icon==='🎃'?'#fff0d8':'#eaf4e7';return 'data:image/svg+xml;charset=UTF-8,'+encodeURIComponent(`<svg xmlns="http://www.w3.org/2000/svg" width="96" height="96"><rect rx="24" width="96" height="96" fill="${hue}"/><text x="48" y="61" font-size="43" text-anchor="middle">${icon}</text></svg>`)}function foodParts(t){if(!t||t==='—')return[];return t.split('＋').map(x=>{let note=(x.match(/（.*?）/)||[''])[0];let raw=x.replace(/（.*?）/g,'').trim();let amount=(raw.match(/(?:约)?\d+(?:\.\d+)?(?:g|mL)/)||[''])[0];let name=raw.replace(/(?:约)?\d+(?:\.\d+)?(?:g|mL)/g,'').replace(/^(冲调后|煮软|去刺|全熟|调稀|瘦|嫩|软烂|压成细碎泥|压成细末)/,'').trim();return{name:name||raw,amount,note,raw}})}function foodCard(x){return `<div class="foodItem"><div class="foodPic"><img src="${svg(x.raw)}" alt="${e(x.name)}配图"></div><div><div class="foodName">${e(x.name)}</div><div class="foodSub">${e(x.amount||'按计划量')}</div></div></div>`}function meal(n,t){if(!t||t==='—')return `<article class="meal"><h2>${n}</h2><div class="food dash">本餐不安排辅食</div></article>`;let parts=foodParts(t);let notes=parts.map(x=>x.note).filter(Boolean).join('；');return `<article class="meal"><h2>${n}</h2><div class="foodGrid">${parts.map(foodCard).join('')}</div><div class="mealFull">${notes?`做法：${e(notes)} · `:''}温热、软烂、湿润，不烫口</div></article>`}function render(){let p=paused(shown),n=ix(shown),row=n>=0&&n<PLAN.length?PLAN[n]:null;$('pick').value=shown;$('start').value=state.start;$('pauseDate').value=shown;let date=shown.replace(/(\d{4})-(\d{2})-(\d{2})/,'$1年$2月$3日');if(shown<state.start){$('hero').innerHTML=`<div class="date">${date}</div><div class="title">辅食尚未开始</div><div class="status">实际开始日期：${state.start}</div>`;$('meals').innerHTML='<div class="help">请选择开始日期之后的日期查看安排。</div>';$('observe').classList.add('hidden');$('advice').innerHTML=''}else if(p){$('hero').innerHTML=`<div class="date">${date}</div><div class="title">今天暂停辅食</div><span class="badge">原计划第 ${n+2} 天</span><div class="status">${e(p.reason||'未填写原因')} · 明天继续同一份计划</div>`;$('meals').innerHTML='<div class="food">今天不安排辅食，继续以奶为主；恢复时不需要加量补回。</div>';$('observe').className='card notice pause';$('observe').innerHTML='<h3>暂停日提醒</h3><p>若因生病、持续腹泻、接种后明显不适而暂停，先以恢复状态为先；准备恢复时从原计划当天继续，不连跳、不补量。</p>';$('advice').innerHTML=''}else if(!row){$('hero').innerHTML=`<div class="date">${date}</div><div class="title">这份计划已完成</div>`;$('meals').innerHTML='<div class="help">请补充下一阶段计划后继续使用。</div>';$('observe').classList.add('hidden');$('advice').innerHTML=''}else{$('hero').innerHTML=`<div class="date">${date}</div><div class="title">今天吃什么</div><span class="badge">计划第 ${n+1} 天 · ${e(row['月龄阶段'])}</span><div class="status">当日总量：${e(row['当日目标总量'])}（是提供目标，不要求吃完）</div>`;$('meals').innerHTML=meal('早餐',row['早餐'])+meal('午餐',row['午餐'])+meal('晚餐',row['晚餐']);let high=/首次|观察|过敏|新增/.test((row['新增食材或观察']||'')+(row['油脂/过敏原/其他']||''));$('observe').className='card notice '+(high?'warn':'');$('observe').classList.remove('hidden');$('observe').innerHTML=`<h3>${high?'今天要重点观察':'今日食材与质地'}</h3><p><b>观察：</b>${e(row['新增食材或观察'])}<br><b>油脂/过敏原：</b>${e(row['油脂/过敏原/其他'])}<br><b>质地：</b>${e(row['质地/进阶'])}</p>`;$('advice').innerHTML=`<div><b>更好吃的小建议</b>${e(row['口感与喂食建议'])}</div><div><b>安全重点</b>${e(row['安全重点'])}</div>`}$('shareHint').textContent=state.shared?'这是家人分享视图。管理员更新暂停后，请重新复制新链接。':'设置只保存在本设备；复制家人链接可同步当前计划状态。';$('pauseList').innerHTML=state.pauses.length?state.pauses.map(x=>`<div class="record"><b>${x.date}</b> · 原计划第 ${ix(x.date)+2} 天<br>${e(x.reason||'未填写原因')}<br><button class="ghost" style="padding:5px 8px;margin-top:5px;font-size:13px" data-del="${x.date}">撤销暂停</button></div>`).join(''):'暂无暂停记录。';document.querySelectorAll('[data-del]').forEach(b=>b.onclick=()=>{if(confirm('撤销该暂停？之后计划会提前一天。')){state.pauses=state.pauses.filter(x=>x.date!==b.dataset.del);state.shared=false;save();render()}})}$('prev').onclick=()=>{shown=plus(shown,-1);render()};$('next').onclick=()=>{shown=plus(shown,1);render()};$('goToday').onclick=()=>{shown=today();render()};$('pick').onchange=x=>{shown=x.target.value;render()};$('openManage').onclick=()=>{$('manager').classList.toggle('hidden')};$('saveStart').onclick=()=>{let d=$('start').value;if(!d)return alert('请选择开始日期');if(!confirm('修改开始日期会重新计算计划，确认？'))return;state.start=d;state.shared=false;save();render()};$('addPause').onclick=()=>{let d=$('pauseDate').value;if(!d||d<state.start)return alert('暂停日期必须不早于开始日期');if(state.pauses.some(x=>x.date===d))return alert('该日期已暂停');if(!confirm(`${d} 暂停并顺延，确认？`))return;state.pauses=[...state.pauses,{date:d,reason:$('pauseReason').value.trim()}].sort((a,b)=>a.date.localeCompare(b.date));state.shared=false;save();shown=d;render()};$('share').onclick=async()=>{let u=new URL(location.href);u.search='';u.searchParams.set('start',state.start);u.searchParams.set('pauses',b64e(JSON.stringify(state.pauses)));try{await navigator.clipboard.writeText(u.href);alert('家人链接已复制。更新暂停后，请重新复制新链接。')}catch(e){prompt('复制此链接：',u.href)}};render();
</script></body></html>'''
OUT_HTML.write_text(html.replace('__PLAN__',plan_json).replace('__RULES__',rules_json),encoding='utf-8')
print(f'Created {OUT_HTML.name} and {OUT_XLSX.name}; {len(revised)} days.')
